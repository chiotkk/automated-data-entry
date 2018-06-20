import openpyxl as xl


wb = xl.load_workbook('output.xlsx')
ws = wb['Invoices']
row = ws.max_row+1


odd_list = ['C', 'E', 'G', 'I', 'K', 'M', 'O', 'Q', 'S', 'U', 'W', 'Y', 'AA', 'AC', 'AE', 'AG', 'AI', 'AK']
even_list = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P', 'R', 'T', 'V', 'X', 'Z', 'AB', 'AD', 'AF', 'AH', 'AJ']


def add_date(date):
    ws['A{}'.format(row)] = int(date)
    wb.save('output.xlsx')


def add_values(values):
    i = 0
    for col in range(2, 38):
        _ = ws.cell(column=col, row=row, value=float(values[i]))
        i += 1
    ws['AL{}'.format(row)] = '=SUM(C{0},E{0},G{0},I{0},K{0},M{0},O{0},Q{0},S{0},U{0},W{0},Y{0},AA{0},AC{0},AE{0},AG{0},AI{0},AK{0})'.format(row)
    
    wb.save('output.xlsx')


def check_values(values, unit_check):
    my_red = xl.styles.colors.Color(rgb='00FF0000')
    my_fill = xl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

    i = 0
    j = 0
    eps = 0.001        
    for col in even_list:
        unit = unit_check[i][:-2]+'.'+unit_check[i][-2:]
        if (abs(float(unit) * float(values[j]) - float(values[j+1])) > eps):
            ws['{}{}'.format(col, row)].fill = my_fill
            ws['{}{}'.format(odd_list[i], row)].fill = my_fill
        i += 1
        j += 2
